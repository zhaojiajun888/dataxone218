# _*_ coding:utf-8 _*_
__author__ = 'jiajun.zhao'

import openpyxl
from openpyxl import load_workbook
from openpyxl import *


class IoExcel():

    # 获取总行数
    def get_row_clo_num(self,sheet_name):
        open_wb = load_workbook("dataexcel.xlsx")
        self.sheet = open_wb[sheet_name]
        rows = self.sheet.max_row
        columns = self.sheet.max_column
        return rows

    def openexcel(self,filepath,sheet_name,starts,ends):
     # 打开excel数据表格
     open_wb = load_workbook(filepath)
     sheet = open_wb[sheet_name]
     data_list = []

     for i in range(starts, ends):
         if i >= starts and i < ends:
             data_dict = {}
             data_dict['casename'] = sheet.cell(i, 1).value
             data_dict['url'] = sheet.cell(i, 2).value
             data_dict['data'] = sheet.cell(i, 3).value
             data_list.append(data_dict)
     return data_list


if __name__ == "__main__":
    excel = IoExcel()
    data = excel.openexcel("../excels/login.xlsx","Sheet1",2,4)
    # num = excel.get_row_clo_num("Sheet1")
    print(data)
